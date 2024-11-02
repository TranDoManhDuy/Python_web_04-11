import rootflaskapp
import datacenter
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import locale
import datetime
# viết các class và function thuộc về product ở đây
# Các class quản lý list chứa Các phương thức xử lý lọc dữ liệu, thêm sửa xóa dữ liệu.
# viết các hàm xử lý và truy vấn, cập nhật dữ liệu SQL tại đây.

# Các route xử lý dữ liệu từ client gửi lên server.
# Các route xử lý dữ liệu từ server trả về client.
class phuongtien():
    def __init__(self) -> None:
        self.__id = ""
        self.__tenphuongtien = ""
        self.__sodangki = ""
        self.__loaiphuongtien = ""
        self.__sochongoi = -1
        self.__tinhtrangxe = ""
        self.__giathue1n = 0
        self.__danhmuc = ""
        self.__ready = "SANSANG"
        
    def setId(self, id: str):
        if len(id) > 0:
            self.__id = id
    def getId(self):
        return self.__id
    def setTenhuongtien(self, tenphuongtien: str):
        if len(tenphuongtien) > 0:
            self.__tenphuongtien = tenphuongtien
    def getTenphuongtien(self):
        return self.__tenphuongtien
    
    def setSodangki(self, sodangki: str):
        if len(sodangki) == 9:
            self.__sodangki = sodangki
    def getSodangki(self):
        return self.__sodangki
    
    def setLoaiphuongtien(self, loaiphuongtien: str):
        self.__loaiphuongtien = loaiphuongtien
    def getLoaiphuongtien(self):
        return self.__loaiphuongtien
    
    def setSochongoi(self, sochongoi: int):
        if sochongoi > 0:
            self.__sochongoi = sochongoi
    def getSochongoi(self):
        return self.__sochongoi
    
    def setTinhtrangxe(self, tinhtrangxe: int):
        if tinhtrangxe >= 0 and tinhtrangxe <= 100:
            self.__tinhtrangxe = tinhtrangxe
    def getTinhtrangxe(self):
        return self.__tinhtrangxe
    
    def setGiathue1n(self, giathue1n: float):
        if giathue1n > 0:
            self.__giathue1n = giathue1n
    def getGiathua1n(self):
        return self.__giathue1n    
    
    def setDanhmuc(self, danhmuc: str):
        if danhmuc != "":
            self.__danhmuc = danhmuc
    def getDanhmuc(self):
        return self.__danhmuc
    
    def setReady(self, ready: str):
        self.__ready = ready
    def getReady(self):
        return self.__ready
    
    def __str__(self) -> str:
        return self.__tenphuongtien + " " + self.__sodangki + " " + self.__loaiphuongtien + " " + str(self.__sochongoi) + " " + self.__tinhtrangxe + " " + str(self.__giathue1n) + " " + self.__danhmuc + " " + self.__ready
    
    # trả về dictionary
    def to_dict(self):
        return {
            "id": self.__id,
            "danhmuc": self.__danhmuc,
            "loaiphuongtien": self.__loaiphuongtien,
            "sodangki": self.__sodangki,
            "tenphuongtien": self.__tenphuongtien,
            "sochongoi": self.__sochongoi,
            "giathue1n": self.__giathue1n,
            "tinhtrangxe": self.__tinhtrangxe,
            "ready": self.__ready
        }
class danhsachphuongtien():
    def __init__(self) -> None:
        self.__danhsachphuongtien = []
    def themphuongtien(self, phuongtien: phuongtien):
        if phuongtien.getSodangki() in [pt.getSodangki() for pt in self.__danhsachphuongtien]:
            return
        self.__danhsachphuongtien.append(phuongtien)
    def layPttheoSDK(self, sodangki):
        for pt in self.__danhsachphuongtien:
            if pt.getSodangki() == sodangki:
                return pt.to_dict()
        return None
    def layPttheoID(self, ID):
        for pt in self.__danhsachphuongtien:
            if pt.getId() == str(ID).rjust(5, "0"):
                return pt
        return None
    def layDSPttheoTen(self, tenphuongtien: str):
        dspt = []
        for pt in self.__danhsachphuongtien:
            if tenphuongtien.lower() in pt.getTenphuongtien().lower():
                dspt.append(pt.to_dict())
        dspt.sort(key=lambda x: x["id"])
        return dspt
    def getlist(self):
        return [pt.to_dict() for pt in self.__danhsachphuongtien]
    def clearList(self):
        self.__danhsachphuongtien.clear()
    def sortList(self):
        self.__danhsachphuongtien.sort(key=lambda x: x.getId())
    
danhsachPT = danhsachphuongtien()
def fetchDataFromDB():
    command = "SELECT * FROM VEHICLE"
    data = datacenter.takedata(command)
    danhsachPT.clearList()
    for i in data:
        pt = phuongtien()
        pt.setId(str(i[0]).rjust(5, "0"))
        pt.setDanhmuc(i[1])
        pt.setLoaiphuongtien(i[2])
        pt.setSodangki(i[3])
        pt.setTenhuongtien(i[4])
        pt.setSochongoi(i[5])
        pt.setGiathue1n(i[6])
        pt.setTinhtrangxe(i[7])
        pt.setReady(i[8])
        danhsachPT.themphuongtien(pt)
# lay toan bo danh sach tu database
@rootflaskapp.app.route("/products_getListPT", methods=["GET"])
def products_getListPT():
    if rootflaskapp.request.method == "GET":
        fetchDataFromDB()
        danhsachPT.sortList()
        return rootflaskapp.jsonify(danhsachPT.getlist())
# Them phuong tien vao database
@rootflaskapp.app.route("/addVehicle", methods=["POST"])
def products_addPT():
    if rootflaskapp.request.method == "POST":
        data = rootflaskapp.request.get_json()
        pt = phuongtien()
        pt.setId(str(int(data["vehicleCode"])))
        pt.setTenhuongtien(data["vehicleName"])
        pt.setSodangki(data["registrationNumber"])
        pt.setLoaiphuongtien(data["vehicleType"])
        pt.setSochongoi(int(data["vehicleSeats"]))
        pt.setTinhtrangxe(int(data["vehicleStatus"]))
        pt.setGiathue1n(float(data["vehiclePrice"]))
        pt.setDanhmuc(data["vehicleCategory"])
        pt.setReady("SANSANG")
        danhsachPT.themphuongtien(pt)
        query = (
            f"INSERT INTO VEHICLE (ID, Category, Type, RegistrationPlate, VName, SeatNumber, Rent, Status, Ready) "
            f"VALUES ('{pt.getId()}', '{pt.getDanhmuc()}', '{pt.getLoaiphuongtien()}', "
            f"'{pt.getSodangki()}', '{pt.getTenphuongtien()}', {pt.getSochongoi()}, "
            f"{pt.getGiathua1n()}, {pt.getTinhtrangxe()}, '{pt.getReady()}')"
        )
        datacenter.pushdata(query)
        return rootflaskapp.jsonify({"status": "success"})
    
# chuyen trang
@rootflaskapp.app.route("/chuyentrangDSPT", methods=["GET"])
def chuyentrangDSPT():
    return rootflaskapp.redirect("/products_list")

# lay phuong tien theo SDK
@rootflaskapp.app.route("/products_getPT", methods=["POST"])
def products_getPT():
    if rootflaskapp.request.method == "POST":
        data = rootflaskapp.request.get_json()
        pt = danhsachPT.layPttheoSDK(data["sodangki"])
        if pt == None:
            return rootflaskapp.jsonify({"status": "fail"})
        return rootflaskapp.jsonify(pt.to_dict())

# Lay danh sach pt theo ten
@rootflaskapp.app.route("/products_getListPTtheoTen", methods=["POST"])
def products_getListPTtheoTen():
    if rootflaskapp.request.method == "POST":
        data = rootflaskapp.request.get_json()
        dspt = danhsachPT.layDSPttheoTen(data["tenphuongtien"])
        return rootflaskapp.jsonify(dspt)

# up date 1 pt
@rootflaskapp.app.route("/products_updatePT", methods=["POST"])
def products_updatePT():
    if rootflaskapp.request.method == "POST":
        data = rootflaskapp.request.get_json()
        pt = danhsachPT.layPttheoSDK(data["sodangki"])
        if pt == None:
            return rootflaskapp.jsonify({"status": "fail"})
        pt.setTenhuongtien(data["tenphuongtien"])
        pt.setLoaiphuongtien(data["loaiphuongtien"])
        pt.setSochongoi(data["sochongoi"])
        pt.setTinhtrangxe(data["tinhtrangxe"])
        pt.setGiathue1n(data["giathue1n"])
        pt.setDanhmuc(data["danhmuc"])
        return rootflaskapp.jsonify({"status": "success"})

# lay ra so thu tu phuong tien tiep theo import vao danh sach
@rootflaskapp.app.route("/getNextVehicleCode", methods=["GET"])
def getNextVehicleCode():
    arr = list(map(lambda x: int(x["id"]), danhsachPT.getlist()))
    if arr.__len__() == 0:
        return rootflaskapp.jsonify({"id": "00001"})
    return rootflaskapp.jsonify({"id": str(max(arr) + 1).rjust(5, "0")})

# kiem tra xem sdk moi, da ton tai trong danh sach sdk chua
@rootflaskapp.app.route("/checkRegistrationNumber", methods=["POST", "GET"])
def checkExistRegistrationNumber():
    if rootflaskapp.request.method == "POST":
        data = rootflaskapp.request.get_json()
        if data["registrationNumber"].isalnum() == False:
            return rootflaskapp.jsonify({"status": "Registration number must be alphanumeric"})
        if data["registrationNumber"].__len__() != 9:
            return rootflaskapp.jsonify({"status": "Length of registration number must be 9 characters"})
        for pt in danhsachPT.getlist():
            if (pt["sodangki"] == data["registrationNumber"]):
                return rootflaskapp.jsonify({"status": "existed registrationNumber"})
    return rootflaskapp.jsonify({"status": "OK"})

# kiem tra 1 so
@rootflaskapp.app.route("/isNumber", methods=["POST"])
def isNumber():
    if rootflaskapp.request.method == "POST":
        data = rootflaskapp.request.get_json()
        if data["number"].isnumeric() == False:
            return rootflaskapp.jsonify({"status": "fail"})
        return rootflaskapp.jsonify({"status": "success"})
    
# fix 
phuongtienSC = phuongtien()
@rootflaskapp.app.route("/getIDForFix", methods=["POST"])
def getIDForFix():
    if rootflaskapp.request.method == "POST":
        data = rootflaskapp.request.get_json()
        pt = danhsachPT.layPttheoID(data["id"]).to_dict()
        global phuongtienSC
        phuongtienSC = pt
        return rootflaskapp.jsonify({"status": "success", "pt": pt})

@rootflaskapp.app.route("/getPT", methods=["GET"])
def getPT():
    return rootflaskapp.jsonify(phuongtienSC)

@rootflaskapp.app.route("/chuyentrang_fixProduct", methods=["GET"])
def chuyentrang_fixProduct():
    return rootflaskapp.redirect(rootflaskapp.url_for("products_edit"))

@rootflaskapp.app.route("/fixVehicle", methods=["POST"])
def fixVehicle():
    if rootflaskapp.request.method == "POST":
        data = rootflaskapp.request.get_json()
        execute = """
            UPDATE VEHICLE
            SET Category = '{}', Type = '{}', RegistrationPlate = '{}', VName = '{}', SeatNumber = {}, Rent = {}, Status = {}
            WHERE ID = '{}'
        """
        datacenter.pushdata(execute.format(data["danhmuc"], data["loaiphuongtien"], data["sodangki"], data["tenphuongtien"], data["sochongoi"], data["giathue1n"], data["tinhtrangxe"], int(data["id"])))
        fetchDataFromDB()
        return rootflaskapp.redirect(rootflaskapp.url_for("products_list"))
# cac ham run khi tai trang
def run_product():
    fetchDataFromDB()
run_product()

# chuyen qua trang them phuong tien
@rootflaskapp.app.route("/chuyentrang_addProduct")
def chuyentrang_addProduct():
    return rootflaskapp.redirect(rootflaskapp.url_for("products_add"))

@rootflaskapp.app.route("/products_list")
def products_list():
    return rootflaskapp.render_template("views/products/list-product.html")

@rootflaskapp.app.route("/products_add")
def products_add():
    return rootflaskapp.render_template("views/products/add-product.html")

@rootflaskapp.app.route("/products_edit")
def products_edit():
    return rootflaskapp.render_template("views/products/edit-product.html")

@rootflaskapp.app.route("/downloadFileExel", methods=['GET', 'POST'])
def downloadFileFunc():
    locale.setlocale(locale.LC_ALL, 'vi_VN.UTF-8')
    if not os.path.exists(rootflaskapp.app.config['UPLOAD_FOLDER']):
        os.makedirs(rootflaskapp.app.config['UPLOAD_FOLDER'])
        
    file_path = os.path.join(rootflaskapp.app.config['UPLOAD_FOLDER'], 'dataProductslist.xlsx')
    if os.path.exists(file_path):
        os.remove(file_path)
    # Tao data, danh sach cac phuong tien
    
    ALL_VEHICLE = datacenter.takedata("SELECT * FROM VEHICLE")
    data = {
        'ID': ["KET XUAT DU LIEU: " + str(datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))] + [str(i[0]).rjust(5, "0") for i in ALL_VEHICLE],
        'Danh muc':[""] + [i[1] for i in ALL_VEHICLE],
        'Loai xe': [""] +[i[2] for i in ALL_VEHICLE],
        'So dang ky': [""] +[i[3] for i in ALL_VEHICLE],
        'Ten xe':[""] + [i[4] for i in ALL_VEHICLE],
        'So cho ngoi': [""] +[i[5] for i in ALL_VEHICLE],
        'Gia thue 1 ngay': [""] +[locale.currency(int(i[6]), grouping=True) for i in ALL_VEHICLE],
        'Tinh trang xe':[""] + [(str(i[7]) + "%") for i in ALL_VEHICLE],
        'Ready': [""] + [i[8] for i in ALL_VEHICLE]
    }
    df = pd.DataFrame(data)
    excel_file = os.path.join(rootflaskapp.app.config['UPLOAD_FOLDER'], 'dataProductslist.xlsx')
    df.to_excel(excel_file, index=False)
    
    wb = load_workbook(excel_file)
    ws = wb.active
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for cell in ws[1]:  # Duyệt qua hàng đầu tiên (tên cột)
        cell.fill = header_fill

    # Lưu file Excel với thay đổi
    wb.save(excel_file)
    try:
        files = os.listdir(rootflaskapp.app.config['UPLOAD_FOLDER'])
        return rootflaskapp.send_from_directory("static", "dataProductslist.xlsx", as_attachment=True)
    except Exception as e:
        rootflaskapp.abort(404)